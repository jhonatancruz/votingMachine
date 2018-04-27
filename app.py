from flask import Flask, render_template, request, session, redirect, url_for
import serial
from openpyxl import Workbook, load_workbook


app= Flask(__name__)
app.secret_key = "any random string"
global ser
# ser= serial.Serial('/dev/ttyACM0', 9600)
ser= serial.Serial('/dev/tty.usbmodem1421', 9600)

# userData=["userinfo","typeofCampaign", "presidentVote", "vicePresidentVote","treasurerVote"]
global cardInfo, checkList
# cardInfo=[["4346DF2B",6,[]],["4ABBDF2B",5,[]],["708CDF2B",4,[]],["86E7DF2B",2,[]],["526EDF2B",1,[]]]
cardInfo= {"4346DF2B":6,"4ABBDF2B":5, "708CDF2B":4, "86E7DF2B":2, "526EDF2B":1 }
cardInfoThroughNumber= {6:"4346DF2B", 5:"4ABBDF2B", 4: "708CDF2B", 2: "86E7DF2B",1:"526EDF2B" }
userData=[]
checkList={"first":[],"last":[],"cardNum":[]}
checkListForVoters={"first":[],"last":[],"cardNum":[]}
votesDirectory={"votesCount":[], "votes":[]}
votes={"barack":1, "donald":2,"bernie":3,"mitt":4,"financial":5,"health":6,"supply":7, "insurance":8, "yes":9, "no":10}



@app.route("/")
def home():
	#scans for admins ids, and gets redirected to startVote
	return render_template("admin.html")

@app.route("/startVote")
def startVote():
	serStr= ser.readline()
	# TODO: if id is in database, then have user choose a campagin, if not then try again
	# usingCard = cardInfo[serStr][1]
	serStr= str(serStr)
	serStr= serStr.strip()
	serStr= serStr[2:-5]
	serStr= serStr.replace(" ","")
	print(serStr)
	if serStr== "D3B865D8":
		return render_template("index.html")
	else:
		return "<h1>Not an Admin ID</h1>"

@app.route("/checkAdmin")
def checkAdmin():
	return render_template("checkAdmin.html")

@app.route("/checking")
def checking():
	serStr= ser.readline()
	# TODO: if id is in database, then have user choose a campagin, if not then try again
	# usingCard = cardInfo[serStr][1]
	serStr= str(serStr)
	serStr= serStr.strip()
	serStr= serStr[2:-5]
	serStr= serStr.replace(" ","")
	print(serStr)
	if serStr== "D3B865D8":
		return redirect(url_for('results'))
	else:
		return '<h1>Not an admin</h1>'

@app.route("/checkid", methods=["POST","GET"])
def checkid():
	#scans for id and whether they have registered
	return render_template("scanner.html")

@app.route("/pickcampaign", methods=["POST", "GET"])
def pickcampaign():
	#should check registered voters excel file to see if they have registered
	serStr= ser.readline()
	serStr= str(serStr)
	serStr= serStr.strip()
	serStr= serStr[2:-5]
	serStr= serStr.replace(" ","")
	print(serStr)
	cardNumber= cardInfo[str(serStr)]
	print(cardNumber)

	wb = Workbook()
	wb = load_workbook('registeredVoters.xlsx')
	ws = wb.active

	#will add all values in excel sheet to dictinary of list for analysis of data
	# No of written Rows in sheet
	r = ws.max_row
	# No of written Columns in sheet
	c = ws.max_column
	# Reading each cell in excel
	for i in range(1, r+1):
	    countTheCols=0
	    for j in range(1, c+1):
	        countTheCols+=1
	        if countTheCols==1:
	            # print("First: ",ws.cell(row=i, column=j).value)
	            checkListForVoters["first"].append(ws.cell(row=i, column=j).value)
	        elif countTheCols==2:
	            # print("Last: ",ws.cell(row=i, column=j).value)
	            checkListForVoters["last"].append(ws.cell(row=i, column=j).value)
	        elif countTheCols==3:
	            # print("CardNum: ",ws.cell(row=i, column=j).value)
	            checkListForVoters["cardNum"].append(int(ws.cell(row=i, column=j).value))

	print(checkListForVoters)

	if cardNumber not in checkListForVoters["cardNum"]:
		return render_template("errorVoting.html")
	else:
		cardNumberLocation= checkListForVoters["cardNum"].index(cardNumber)
		print("location of card number in list",cardNumberLocation)
		userInfo= checkListForVoters["first"][cardNumberLocation]+" "+checkListForVoters["last"][cardNumberLocation]
		print("Name of Card Owner:", userInfo)
		# return render_template("pickcampaign.html", userInfo=userInfo, cardNumber= cardNumber)
		return render_template("newhtml.html", userInfo=userInfo, cardNumber=cardNumber)

@app.route("/register", methods=["GET", "POST"])
def register():
	#registration form
	return render_template("registration.html")

@app.route("/index", methods=["GET", "POST"])
def index():
	return render_template("index.html")

@app.route("/saveinfo",methods=["POST"])
def saveinfo():
    # saves the registration data
	if request.method == "POST":
		firstName= request.form["firstName"]
		lastName= request.form["lastName"]
		cardNumber= request.form["idnumber"]

		#for registration, adds the next value in the new row
		wb = Workbook()
		wb = load_workbook('registeredVoters.xlsx')
		ws = wb.active

		#will add all values in excel sheet to dictinary of list for analysis of data
		# No of written Rows in sheet
		r = ws.max_row
		# No of written Columns in sheet
		c = ws.max_column
		# Reading each cell in excel
		for i in range(1, r+1):
		    countTheCols=0
		    for j in range(1, c+1):
		        countTheCols+=1
		        if countTheCols==1:
		            # print("First: ",ws.cell(row=i, column=j).value)
		            checkList["first"].append(ws.cell(row=i, column=j).value)
		        elif countTheCols==2:
		            # print("Last: ",ws.cell(row=i, column=j).value)
		            checkList["last"].append(ws.cell(row=i, column=j).value)
		        elif countTheCols==3:
		            # print("CardNum: ",ws.cell(row=i, column=j).value)
		            checkList["cardNum"].append(ws.cell(row=i, column=j).value)

		print(checkList)

		#will analyze data and check if card number has been registers, and if someone with the same name has register
		if cardNumber in checkList["cardNum"]:
			print("no sorry someone is already registered with this card")
			return "<h1>no sorry someone is already registered with this card</h1>"
		# elif firstName in checkList["first"]:
		#     locationOfMatchVal= checkList["first"].index(firstName)
		#     print("yea there... checking for last name at index", checkList["first"].index(firstName), " since thats where we found first name")
		#     if checkList["last"][locationOfMatchVal] == lastName:
				# return "<h1>Found someone with the exact name, sorry</h1>"
		else:
		    print("not there, you are able to register")

		    #will add data to next open row
		    max_row=ws.max_row #find the max row that data is in
		    print(max_row)
		    newRow= max_row+1
		    ws.cell(row=newRow, column=1).value= firstName
		    ws.cell(row=newRow, column=2).value= lastName
		    ws.cell(row=newRow, column=3).value= cardNumber
			# wb.save('registeredVoters.xlsx')
			# return '<h1>You were able to register</h1>'


		wb.save('registeredVoters.xlsx')
		return redirect(url_for("home"))
	else:
		return render_template("registration.html")


@app.route("/voting", methods=["POST"])
def voting():
	#based on the choosen campagin it will give you a certain campagin
	if request.method == "POST":
		president= request.form['president']
		blockchain= request.form['blockchain']
		texas= request.form['texas']

		wb = Workbook()
		wb = load_workbook('votes.xlsx')
		ws = wb.active

		#will add all values in excel sheet to dictinary of list for analysis of data
		# No of written Rows in sheet
		r = ws.max_row
		# No of written Columns in sheet
		c = ws.max_column
		# Reading each cell in excel
		for i in range(1, r+1):
		    countTheCols=0
		    for j in range(1, c+1):
		        countTheCols+=1
		        if countTheCols==1:
		            # print("First: ",ws.cell(row=i, column=j).value)
		            votesDirectory["votes"].append(ws.cell(row=i, column=j).value)
		        elif countTheCols==2:
		            # print("Last: ",ws.cell(row=i, column=j).value)
		            votesDirectory["votesCount"].append(ws.cell(row=i, column=j).value)




		print(votesDirectory)
		print(president, blockchain, texas)

		locationOfPres= votesDirectory["votes"].index(president)
		presVotes= votesDirectory["votesCount"][locationOfPres]
		presVotesUpdated= int(presVotes)+1

		locationOfBlock= votesDirectory["votes"].index(blockchain)
		blockVotes= votesDirectory["votesCount"][locationOfBlock]
		blockVotesUpdated= int(blockVotes)+1

		locationOfTexas= votesDirectory["votes"].index(texas)
		texasVotes= votesDirectory["votesCount"][locationOfTexas]
		texasVotesUpdated= int(texasVotes)+1

		ws.cell(row=votes[president], column=2).value= presVotesUpdated
		ws.cell(row=votes[blockchain], column=2).value= blockVotesUpdated
		ws.cell(row=votes[texas], column=2).value= texasVotesUpdated
        #
		# wb.save('registeredVoters.xlsx')
		# return redirect(url_for("home"))

		wb.save('votes.xlsx')
		return render_template("thankYou.html")


@app.route("/results", methods=["POST", "GET"])
def results():
	wb = Workbook()
	wb = load_workbook('votes.xlsx')
	ws = wb.active

	#will add all values in excel sheet to dictinary of list for analysis of data
	# No of written Rows in sheet
	r = ws.max_row
	# No of written Columns in sheet
	c = ws.max_column
	# Reading each cell in excel
	for i in range(1, r+1):
	    countTheCols=0
	    for j in range(1, c+1):
	        countTheCols+=1
	        if countTheCols==1:
	            # print("First: ",ws.cell(row=i, column=j).value)
	            votesDirectory["votes"].append(ws.cell(row=i, column=j).value)
	        elif countTheCols==2:
	            # print("Last: ",ws.cell(row=i, column=j).value)
	            votesDirectory["votesCount"].append(ws.cell(row=i, column=j).value)
	voters= votesDirectory["votes"]
	votes= votesDirectory["votesCount"]
	allVotes= votesDirectory
	return render_template("ballot.html", votes=votes, voters=voters, allVotes=allVotes)





# @app.route("/vicePresVote", methods=["POST"])
# def vicePresVote():
# 	if request.method == "POST":
# 		presVote= request.form["optradio"]
# 		userData.append(presVote)
# 		print(presVote)
# 		return render_template("vicepresVote.html")
# 	else:
# 		print("didnt record a president vote")
# 		return render_template("presidentialVote.html")
#
# @app.route("/treasurerVote", methods=["POST"])
# def treasurerVote():
# 	if request.method == "POST":
# 		vicePresVote= request.form["optradio"]
# 		userData.append(vicePresVote)
# 		print(vicePresVote)
# 		return render_template("treasurerVote.html")
# 	else:
# 		print("didnt record a vice president vote")
# 		return render_template("vicePresVote.html")

@app.route("/thankYou", methods=["POST"])
def thankYou():
	if request.method == "POST":
		treasurerVote= request.form["optradio"]
		userData.append(treasurerVote)
		print(treasurerVote)
		#~ session[""]
		return render_template("thankYou.html", userData=userData)
	else:
		print("didnt record a treasurer vote")
		return render_template("treasurerVote.html")


if __name__=="__main__":
	app.run(debug=True)
