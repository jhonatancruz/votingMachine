from openpyxl import Workbook, load_workbook

firstName= 'Stevens'
lastName= 'Cruz'
cardNumber= 6
checkList={"first":[],"last":[],"cardNum":[]}

#for registration, adds the next value in the new row
wb = Workbook()
wb = load_workbook('registeredVoters.xlsx')
ws = wb.active

#will add all values in excel sheet to dictinary of list for analysis of data
# No of written Rows in sheet
r = ws.max_row
# No of written Columns in sheet
c = ws.max_column
# Reading each cell in excel
for i in range(1, r+1):
    countTheCols=0
    for j in range(1, c+1):
        countTheCols+=1
        if countTheCols==1:
            # print("First: ",ws.cell(row=i, column=j).value)
            checkList["first"].append(ws.cell(row=i, column=j).value)
        elif countTheCols==2:
            # print("Last: ",ws.cell(row=i, column=j).value)
            checkList["last"].append(ws.cell(row=i, column=j).value)
        elif countTheCols==3:
            # print("CardNum: ",ws.cell(row=i, column=j).value)
            checkList["cardNum"].append(ws.cell(row=i, column=j).value)

print(checkList)

#will analyze data and check if card number has been registers, and if someone with the same name has register
if cardNumber in checkList["cardNum"]:
    print("no sorry someone is already registered with this card")
elif firstName in checkList["first"]:
    locationOfMatchVal= checkList["first"].index(firstName)
    print("yea there... checking for last name at index", checkList["first"].index(firstName), " since thats where we found first name")
    if checkList["last"][locationOfMatchVal] == lastName:
        print("no sorry, we found someone with that exact name")
else:
    print("not there, you are able to register")
    #will add data to next open row
    max_row=ws.max_row #find the max row that data is in
    print(max_row)
    newRow= max_row+1
    ws.cell(row=newRow, column=1).value= firstName
    ws.cell(row=newRow, column=2).value= lastName
    ws.cell(row=newRow, column=3).value= cardNumber



wb.save('registeredVoters.xlsx')
